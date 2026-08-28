"""Leitura da lista de produtos e geracao do relatorio Excel do RPA."""

from __future__ import annotations

import re
import unicodedata
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .errors import SpreadsheetError
from .models import CollectionRow, ProductInput


HEADER_ALIASES = {
    "produto": {
        "produto",
        "nome",
        "nomeproduto",
        "item",
        "descricao",
        "descricaoproduto",
    },
    "marca": {"marca", "fornecedor", "fabricante"},
    "modelo": {"modelo"},
    "sku": {
        "sku",
        "material",
        "codigo",
        "codigoproduto",
        "codigodoproduto",
        "coditem",
    },
    "quantidade": {"quantidade", "qtd", "qtde", "quantidadesolicitada"},
}

RESULT_HEADERS = [
    "Linha origem",
    "Produto solicitado",
    "Marca solicitada",
    "Modelo",
    "SKU",
    "Quantidade solicitada",
    "Consulta",
    "Fonte",
    "Posicao",
    "Nome produto encontrado",
    "Marca encontrada",
    "Quantidade/embalagem encontrada",
    "Preco",
    "Preco maximo",
    "Classificacao",
    "Similaridade (%)",
    "Possivel produto parecido",
    "Loja",
    "Link de compra",
    "Imagem",
    "Coletado em UTC",
    "Status",
    "Erro",
]

SIMILAR_HEADERS = [
    "Produto solicitado",
    "Quantidade solicitada",
    "Fonte",
    "Nome produto parecido",
    "Marca encontrada",
    "Quantidade/embalagem",
    "Preco",
    "Similaridade (%)",
    "Loja",
    "Link de compra",
]

HEADER_FILL = PatternFill("solid", fgColor="183B56")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ERROR_FILL = PatternFill("solid", fgColor="FCE8E6")
SIMILAR_FILL = PatternFill("solid", fgColor="FFF2CC")
DIVERGENT_FILL = PatternFill("solid", fgColor="E7E6E6")


def _normalize_header(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]", "", text.casefold())


def _cell_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def read_products(
    path: str | Path,
    sheet_name: str | None = None,
    max_products: int = 1000,
) -> list[ProductInput]:
    source = Path(path)
    if source.suffix.casefold() != ".xlsx":
        raise SpreadsheetError("A entrada deve ser um arquivo .xlsx")
    if not source.is_file():
        raise SpreadsheetError(f"Planilha nao encontrada: {source}")
    if max_products < 1:
        raise SpreadsheetError("max_products deve ser maior que zero")

    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except Exception as exc:
        raise SpreadsheetError("Nao foi possivel abrir a planilha") from exc

    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise SpreadsheetError(f"Aba nao encontrada: {sheet_name}")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        header_row = _find_header_row(sheet)
        if header_row is None:
            first_row = next(
                sheet.iter_rows(min_row=1, max_row=1, values_only=True), None
            )
            if not first_row or not any(_cell_text(value) for value in first_row):
                raise SpreadsheetError("A planilha esta vazia")
            raise SpreadsheetError("Coluna obrigatoria 'Produto' nao encontrada")
        header_values = next(
            sheet.iter_rows(
                min_row=header_row, max_row=header_row, values_only=True
            ),
            None,
        )
        if not header_values:
            raise SpreadsheetError("A planilha esta vazia")
        normalized = [_normalize_header(value) for value in header_values]
        indexes: dict[str, int] = {}
        for canonical, aliases in HEADER_ALIASES.items():
            for index, header in enumerate(normalized):
                if header in aliases:
                    indexes[canonical] = index
                    break
        if "produto" not in indexes:
            raise SpreadsheetError("Coluna obrigatoria 'Produto' nao encontrada")

        products: list[ProductInput] = []
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            produto = _value_at(values, indexes.get("produto"))
            marca = _value_at(values, indexes.get("marca"))
            modelo = _value_at(values, indexes.get("modelo"))
            sku = _value_at(values, indexes.get("sku"))
            quantidade = _value_at(values, indexes.get("quantidade"))
            if not any((produto, marca, modelo, sku, quantidade)):
                continue
            if not produto:
                raise SpreadsheetError(f"Linha {row_number}: Produto esta vazio")
            products.append(
                ProductInput(row_number, produto, marca, modelo, sku, quantidade)
            )
            if len(products) > max_products:
                raise SpreadsheetError(
                    f"A planilha excede o limite de {max_products} produtos"
                )
        if not products:
            raise SpreadsheetError(
                f"A planilha '{source.resolve()}' nao possui produtos preenchidos. "
                "Digite um nome na coluna 'Produto' a partir da linha 2, salve o "
                "arquivo e execute novamente."
            )
        return products
    finally:
        workbook.close()


def _value_at(values: tuple[object, ...], index: int | None) -> str | None:
    if index is None or index >= len(values):
        return None
    return _cell_text(values[index])


def _find_header_row(sheet) -> int | None:
    for row_number, values in enumerate(
        sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1
    ):
        normalized = {_normalize_header(value) for value in values}
        if normalized & HEADER_ALIASES["produto"] and normalized & HEADER_ALIASES["sku"]:
            return row_number
        if "produto" in normalized:
            return row_number
    return None


def create_template(path: str | Path) -> Path:
    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Produtos"
    headers = ["Produto", "Marca", "Modelo", "SKU", "Quantidade"]
    sheet.append(headers)
    _style_header(sheet)
    notes = {
        "A1": "Obrigatorio. Nome do produto pesquisado pelo RPA.",
        "B1": "Opcional. Melhora a busca e a validacao de similaridade.",
        "C1": "Opcional. Modelo ou especificacao principal.",
        "D1": "Opcional. Codigo interno ou do fabricante.",
        "E1": "Opcional. Quantidade que deve ser comprada; nao e estoque do anuncio.",
    }
    for coordinate, text in notes.items():
        sheet[coordinate].comment = Comment(text, "Ronaldo Bueno")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:E1"
    for index, width in enumerate([42, 22, 24, 22, 18], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    workbook.save(destination)
    workbook.close()
    return destination


def export_results(rows: list[CollectionRow], path: str | Path) -> Path:
    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    results_sheet = workbook.active
    results_sheet.title = "Resultados"
    results_sheet.append(RESULT_HEADERS)

    for row in rows:
        item = row.result
        results_sheet.append(
            [
                row.product.row_number,
                row.product.produto,
                row.product.marca,
                row.product.modelo,
                row.product.sku,
                row.product.quantidade_solicitada,
                row.product.query,
                row.provider,
                item.rank if item else None,
                item.title if item else None,
                item.brand if item else None,
                item.package_quantity if item else None,
                item.price_min if item else None,
                item.price_max if item else None,
                item.match_type if item else None,
                item.similarity_score / 100 if item else None,
                "Sim" if item and item.possible_similar else "Nao",
                item.seller if item else None,
                item.purchase_url if item else None,
                item.image_url if item else None,
                row.collected_at.replace(microsecond=0).isoformat(),
                row.status,
                row.error,
            ]
        )

    _format_results_sheet(results_sheet)
    _build_similar_sheet(workbook, rows)
    _build_summary_sheet(workbook, rows)
    workbook.save(destination)
    workbook.close()
    return destination


def _style_header(sheet) -> None:
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _add_table(sheet, name: str) -> None:
    if sheet.max_row < 2:
        return
    table = Table(displayName=name, ref=sheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def _format_results_sheet(sheet) -> None:
    _style_header(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 36
    widths = [
        12, 30, 20, 20, 18, 20, 38, 18, 10, 44, 20, 24, 16, 16, 18, 18,
        22, 22, 50, 44, 24, 18, 48,
    ]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row_index in range(2, sheet.max_row + 1):
        sheet.cell(row_index, 13).number_format = 'R$ #,##0.00'
        sheet.cell(row_index, 14).number_format = 'R$ #,##0.00'
        sheet.cell(row_index, 16).number_format = "0.0%"
        for column in (19, 20):
            cell = sheet.cell(row_index, column)
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"
        match_type = sheet.cell(row_index, 15).value
        if match_type == "SIMILAR":
            for cell in sheet[row_index]:
                cell.fill = SIMILAR_FILL
        elif match_type == "DIVERGENTE":
            for cell in sheet[row_index]:
                cell.fill = DIVERGENT_FILL
        if sheet.cell(row_index, 22).value == "ERRO":
            for cell in sheet[row_index]:
                cell.fill = ERROR_FILL
    _add_table(sheet, "ResultadosRpa")


def _build_similar_sheet(workbook: Workbook, rows: list[CollectionRow]) -> None:
    sheet = workbook.create_sheet("Produtos similares")
    sheet.append(SIMILAR_HEADERS)
    for row in rows:
        item = row.result
        if item is None or not item.possible_similar:
            continue
        sheet.append(
            [
                row.product.produto,
                row.product.quantidade_solicitada,
                row.provider,
                item.title,
                item.brand,
                item.package_quantity,
                item.price_min,
                item.similarity_score / 100,
                item.seller,
                item.purchase_url,
            ]
        )
    _style_header(sheet)
    sheet.freeze_panes = "A2"
    for index, width in enumerate(
        [30, 20, 18, 44, 20, 24, 16, 18, 22, 50], start=1
    ):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row_index in range(2, sheet.max_row + 1):
        sheet.cell(row_index, 7).number_format = 'R$ #,##0.00'
        sheet.cell(row_index, 8).number_format = "0.0%"
        link_cell = sheet.cell(row_index, 10)
        if link_cell.value:
            link_cell.hyperlink = str(link_cell.value)
            link_cell.style = "Hyperlink"
    _add_table(sheet, "ProdutosSimilares")


def _build_summary_sheet(workbook: Workbook, rows: list[CollectionRow]) -> None:
    sheet = workbook.create_sheet("Resumo")
    sheet.append(
        [
            "Produto",
            "Marca solicitada",
            "SKU",
            "Quantidade solicitada",
            "Ofertas encontradas",
            "Compativeis",
            "Similares",
            "Menor preco",
            "Fonte menor preco",
            "Link de compra",
        ]
    )
    grouped: dict[tuple[str, str | None], list[CollectionRow]] = defaultdict(list)
    for row in rows:
        grouped[(row.product.produto, row.product.sku)].append(row)

    for (product_name, sku), product_rows in grouped.items():
        successes = [
            row
            for row in product_rows
            if row.result is not None and row.result.price_min is not None
        ]
        cheapest = min(successes, key=lambda row: row.result.price_min) if successes else None
        result = cheapest.result if cheapest else None
        product = product_rows[0].product
        sheet.append(
            [
                product_name,
                product.marca,
                sku,
                product.quantidade_solicitada,
                sum(1 for row in product_rows if row.status == "OK"),
                sum(
                    1
                    for row in product_rows
                    if row.result is not None and row.result.match_type == "COMPATIVEL"
                ),
                sum(
                    1
                    for row in product_rows
                    if row.result is not None and row.result.possible_similar
                ),
                result.price_min if result else None,
                cheapest.provider if cheapest else None,
                result.purchase_url if result else None,
            ]
        )

    _style_header(sheet)
    sheet.freeze_panes = "A2"
    for index, width in enumerate(
        [34, 20, 18, 20, 20, 14, 14, 18, 22, 50], start=1
    ):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row_index in range(2, sheet.max_row + 1):
        sheet.cell(row_index, 8).number_format = 'R$ #,##0.00'
        link_cell = sheet.cell(row_index, 10)
        if link_cell.value:
            link_cell.hyperlink = str(link_cell.value)
            link_cell.style = "Hyperlink"
    _add_table(sheet, "ResumoRpa")
