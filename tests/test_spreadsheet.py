import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from copercitrus_price_collector.errors import SpreadsheetError
from copercitrus_price_collector.models import CollectionRow, ProductInput, SearchResult
from copercitrus_price_collector.spreadsheet import create_template, export_results, read_products


class SpreadsheetTest(unittest.TestCase):
    def test_reads_quantity_and_builds_query_without_purchase_quantity(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "produtos.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Nome", "Marca", "Modelo", "Codigo do produto", "Qtd"])
            sheet.append(["Notebook", "Dell", "Inspiron 15", "ABC-1", 25])
            workbook.save(path)
            workbook.close()

            products = read_products(path)

            self.assertEqual("Notebook Dell Inspiron 15 ABC-1", products[0].query)
            self.assertEqual("25", products[0].quantidade_solicitada)

    def test_rejects_missing_product_header(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "produtos.xlsx"
            workbook = Workbook()
            workbook.active.append(["Marca", "Modelo"])
            workbook.active.append(["Dell", "X"])
            workbook.save(path)
            workbook.close()

            with self.assertRaisesRegex(SpreadsheetError, "Produto"):
                read_products(path)

    def test_creates_template_and_three_output_sheets(self):
        with tempfile.TemporaryDirectory() as directory:
            template_path = create_template(Path(directory) / "modelo.xlsx")
            template = load_workbook(template_path)
            self.assertEqual(
                ["Produto", "Marca", "Modelo", "SKU", "Quantidade"],
                [cell.value for cell in template["Produtos"][1]],
            )
            template.close()

            product = ProductInput(2, "Mouse", "Logitech", "M170", "SKU-1", "10")
            result = SearchResult(
                provider="Shopee",
                rank=1,
                title="Mouse Logitech M185 kit 2 unidades",
                description="Mouse sem fio",
                price_min=79.9,
                price_max=79.9,
                currency="BRL",
                purchase_url="https://shopee.example/item",
                brand="Logitech",
                package_quantity="2 un",
                similarity_score=67.5,
                match_type="SIMILAR",
                seller="Loja",
            )
            output = export_results(
                [CollectionRow.success(product, result)],
                Path(directory) / "resultado.xlsx",
            )
            workbook = load_workbook(output)
            self.assertEqual(
                ["Resultados", "Produtos similares", "Resumo"], workbook.sheetnames
            )
            self.assertEqual("OK", workbook["Resultados"]["V2"].value)
            self.assertEqual("Sim", workbook["Resultados"]["Q2"].value)
            self.assertEqual("Mouse Logitech M185 kit 2 unidades", workbook["Produtos similares"]["D2"].value)
            self.assertEqual(79.9, workbook["Resumo"]["H2"].value)
            self.assertEqual(
                "https://shopee.example/item",
                workbook["Resumo"]["J2"].hyperlink.target,
            )
            workbook.close()


if __name__ == "__main__":
    unittest.main()
